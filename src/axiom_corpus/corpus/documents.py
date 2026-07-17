"""Generic official-document ingestion for policy sources."""

from __future__ import annotations

import base64
import re
import shutil
import subprocess
import sys
import tempfile
import time
import warnings
import zipfile
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from json import dumps as json_dumps
from json import loads as json_loads
from pathlib import Path
from typing import Any, Self, TextIO, cast
from xml.etree import ElementTree

import fitz
import requests
import xlrd
import yaml
from bs4 import BeautifulSoup, FeatureNotFound
from bs4.element import Comment, Tag
from openpyxl import load_workbook
from urllib3.exceptions import InsecureRequestWarning

from axiom_corpus.corpus.artifacts import CorpusArtifactStore, safe_segment
from axiom_corpus.corpus.coverage import ProvisionCoverageReport, compare_provision_coverage
from axiom_corpus.corpus.models import DocumentClass, ProvisionRecord, SourceInventoryItem
from axiom_corpus.corpus.supabase import deterministic_provision_id

OFFICIAL_DOCUMENT_USER_AGENT = (
    "Axiom/1.0 (Legal Archive; contact@axiom-foundation.org) "
    "https://github.com/TheAxiomFoundation/axiom-corpus"
)
OFFICIAL_DOCUMENT_BROWSER_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
OFFICIAL_DOCUMENT_BROWSER_IMPERSONATION = "chrome120"
_BROWSER_FALLBACK_STATUSES = {403, 404, 406}
_REQUEST_RETRY_STATUSES = {429, 500, 502, 503, 504}
_REQUEST_RETRY_ATTEMPTS = 4
_REQUEST_RETRY_BASE_DELAY_SECONDS = 0.5
_RANGE_FETCH_CHUNK_SIZE_BYTES = 1024 * 1024
_CONTENT_RANGE_RE = re.compile(r"^bytes (?P<start>\d+)-(?P<end>\d+)/(?P<total>\d+|\*)$")
_GOOGLE_DRIVE_FILE_RE = re.compile(r"https?://drive\.google\.com/file/d/([^/]+)/")
_MAPBOX_PUBLIC_TOKEN_RE = re.compile(rb"pk\.[A-Za-z0-9_-]+\.[A-Za-z0-9_-]+")
_MAPBOX_PUBLIC_TOKEN_PLACEHOLDER = b"[redacted-mapbox-public-token]"
_LEGACY_WORD_DOCUMENT_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_HEADING_TAGS = {"h1", "h2", "h3", "h4", "h5", "h6"}
_TEXT_TAGS = _HEADING_TAGS | {"p", "li", "table", "blockquote"}
_NON_HEADING_UPPERCASE_LINES = {
    "HANDBOOK BEGINS HERE",
    "HANDBOOK CONTINUES",
    "HANDBOOK CONTINUE",
    "HANDBOOK ENDS HERE",
}
_WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


@dataclass(frozen=True)
class OfficialDocumentSource:
    """One primary official document to snapshot and normalize."""

    source_id: str
    jurisdiction: str
    document_class: str
    title: str
    source_url: str
    citation_path: str | None = None
    download_url: str | None = None
    source_format: str | None = None
    source_as_of: str | None = None
    expression_date: str | None = None
    language: str | None = None
    local_path: str | None = None
    request: dict[str, Any] | None = None
    extraction: dict[str, Any] | None = None
    metadata: dict[str, Any] | None = None

    @classmethod
    def from_mapping(cls, data: dict[str, Any]) -> Self:
        request = data.get("request")
        if request is not None and not isinstance(request, dict):
            raise ValueError("official document request config must be a mapping")
        extraction = data.get("extraction")
        if extraction is not None and not isinstance(extraction, dict):
            raise ValueError("official document extraction config must be a mapping")
        language = data.get("language")
        if isinstance(language, bool):
            raise ValueError(
                "official document language must be a string; YAML 1.1 parses "
                'unquoted codes like `language: no` as booleans - quote it ("no")'
            )
        return cls(
            source_id=str(data["source_id"]),
            jurisdiction=str(data["jurisdiction"]),
            document_class=str(data.get("document_class", DocumentClass.POLICY.value)),
            title=str(data["title"]),
            source_url=str(data["source_url"]),
            citation_path=data.get("citation_path"),
            download_url=data.get("download_url"),
            source_format=data.get("source_format"),
            source_as_of=data.get("source_as_of"),
            expression_date=data.get("expression_date"),
            language=str(language) if language is not None else None,
            local_path=data.get("local_path"),
            request=request,
            extraction=extraction,
            metadata=data.get("metadata"),
        )


@dataclass(frozen=True)
class OfficialDocumentManifest:
    """Manifest of primary official documents for one corpus scope."""

    documents: tuple[OfficialDocumentSource, ...]

    @classmethod
    def load(cls, path: str | Path) -> Self:
        data = yaml.safe_load(Path(path).read_text())
        if not isinstance(data, dict):
            raise ValueError("official document manifest must be a YAML mapping")
        documents = data.get("documents")
        if not isinstance(documents, list):
            raise ValueError("official document manifest must contain a documents list")
        return cls(
            documents=tuple(
                OfficialDocumentSource.from_mapping(row)
                for row in documents
                if isinstance(row, dict)
            )
        )

    def require_unique_sources(self) -> None:
        seen: set[str] = set()
        for source in self.documents:
            if source.source_id in seen:
                raise ValueError(f"duplicate source_id: {source.source_id}")
            seen.add(source.source_id)


@dataclass(frozen=True)
class OfficialDocumentExtractReport:
    """Result from a generic official-document extraction run."""

    jurisdiction: str
    document_class: str
    document_count: int
    block_count: int
    provisions_written: int
    inventory_path: Path
    provisions_path: Path
    coverage_path: Path
    coverage: ProvisionCoverageReport
    source_paths: tuple[Path, ...]


@dataclass(frozen=True)
class _DownloadedDocument:
    source: OfficialDocumentSource
    content: bytes
    content_type: str | None
    final_url: str


@dataclass(frozen=True)
class _DocumentBlock:
    kind: str
    ordinal: int
    heading: str | None
    body: str
    metadata: dict[str, Any]


def official_documents_run_id(
    version: str,
    *,
    only_source_id: str | None = None,
    limit: int | None = None,
) -> str:
    """Return a scoped run id for a manifest-driven official-document run."""
    parts = [version]
    if only_source_id:
        parts.append(safe_segment(only_source_id))
    if limit is not None:
        parts.append(f"limit-{limit}")
    return "-".join(parts)


def extract_official_documents(
    store: CorpusArtifactStore,
    *,
    manifest_path: str | Path,
    version: str,
    source_as_of: str | None = None,
    expression_date: date | str | None = None,
    only_source_id: str | None = None,
    limit: int | None = None,
    progress_stream: TextIO | None = None,
) -> OfficialDocumentExtractReport:
    """Snapshot official HTML/PDF documents and extract normalized records."""
    manifest = OfficialDocumentManifest.load(manifest_path)
    manifest.require_unique_sources()
    documents = _select_documents(manifest.documents, only_source_id=only_source_id, limit=limit)
    if not documents:
        raise ValueError("no official documents selected")
    jurisdiction, document_class = _single_scope(documents)
    run_id = official_documents_run_id(version, only_source_id=only_source_id, limit=limit)
    default_source_as_of = source_as_of or version
    default_expression_date = _date_text(expression_date, default_source_as_of)

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": OFFICIAL_DOCUMENT_USER_AGENT,
            "Accept": "text/html,application/pdf,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )

    source_paths: list[Path] = []
    inventory: list[SourceInventoryItem] = []
    records: list[ProvisionRecord] = []
    block_count = 0

    for source in documents:
        _progress(progress_stream, f"extracting {source.source_id}")
        downloaded = _download_document(source, session=session)
        source_format = _infer_source_format(source, downloaded)
        content = _sanitize_official_document_content(downloaded.content, source_format)
        relative_source = (
            f"official-documents/{safe_segment(source.source_id)}{_extension(source_format)}"
        )
        artifact_path = store.source_path(jurisdiction, document_class, run_id, relative_source)
        source_sha = store.write_bytes(artifact_path, content)
        source_paths.append(artifact_path)
        source_key = f"sources/{jurisdiction}/{document_class}/{run_id}/{relative_source}"
        source_as_of_text = source.source_as_of or default_source_as_of
        expression_date_text = source.expression_date or default_expression_date

        blocks = tuple(
            _extract_blocks(
                content,
                source_format,
                source_url=source.source_url,
                title=source.title,
                extraction=source.extraction,
            )
        )
        block_count += len(blocks)
        inventory.extend(
            _inventory_items(
                source,
                blocks=blocks,
                source_key=source_key,
                source_format=source_format,
                source_sha=source_sha,
                content_type=downloaded.content_type,
                final_url=downloaded.final_url,
            )
        )
        records.extend(
            _provision_records(
                source,
                blocks=blocks,
                version=run_id,
                source_key=source_key,
                source_format=source_format,
                source_as_of=source_as_of_text,
                expression_date=expression_date_text,
                content_type=downloaded.content_type,
                final_url=downloaded.final_url,
            )
        )

    inventory_path = store.inventory_path(jurisdiction, document_class, run_id)
    provisions_path = store.provisions_path(jurisdiction, document_class, run_id)
    coverage_path = store.coverage_path(jurisdiction, document_class, run_id)
    store.write_inventory(inventory_path, inventory)
    store.write_provisions(provisions_path, records)
    coverage = compare_provision_coverage(
        tuple(inventory),
        tuple(records),
        jurisdiction=jurisdiction,
        document_class=document_class,
        version=run_id,
    )
    store.write_json(coverage_path, coverage.to_mapping())

    return OfficialDocumentExtractReport(
        jurisdiction=jurisdiction,
        document_class=document_class,
        document_count=len(documents),
        block_count=block_count,
        provisions_written=len(records),
        inventory_path=inventory_path,
        provisions_path=provisions_path,
        coverage_path=coverage_path,
        coverage=coverage,
        source_paths=tuple(source_paths),
    )


def google_drive_download_url(url: str) -> str | None:
    """Return a direct download URL for a public Google Drive file URL."""
    match = _GOOGLE_DRIVE_FILE_RE.match(url)
    if not match:
        return None
    return f"https://drive.google.com/uc?export=download&id={match.group(1)}"


def _select_documents(
    documents: tuple[OfficialDocumentSource, ...],
    *,
    only_source_id: str | None,
    limit: int | None,
) -> tuple[OfficialDocumentSource, ...]:
    selected = [source for source in documents if only_source_id in {None, source.source_id}]
    if limit is not None:
        selected = selected[:limit]
    return tuple(selected)


def _single_scope(documents: tuple[OfficialDocumentSource, ...]) -> tuple[str, str]:
    jurisdictions = {source.jurisdiction for source in documents}
    document_classes = {source.document_class for source in documents}
    if len(jurisdictions) != 1 or len(document_classes) != 1:
        raise ValueError("official document extraction requires one jurisdiction/document_class")
    return next(iter(jurisdictions)), next(iter(document_classes))


def _download_document(
    source: OfficialDocumentSource,
    *,
    session: requests.Session,
) -> _DownloadedDocument:
    if source.local_path:
        path = Path(source.local_path)
        return _DownloadedDocument(
            source=source,
            content=path.read_bytes(),
            content_type=None,
            final_url=path.as_uri(),
        )
    download_url = (
        source.download_url or google_drive_download_url(source.source_url) or source.source_url
    )
    request_config = source.request or {}
    verify = bool(request_config.get("verify_tls", True))
    request_headers = _request_headers_from_config(request_config)
    if request_config.get("range_fetch"):
        if request_config.get("range_backend") == "curl":
            return _download_document_by_curl_ranges(
                source,
                download_url,
                headers=request_headers,
                verify=verify,
                chunk_size=int(
                    request_config.get("range_chunk_size", _RANGE_FETCH_CHUNK_SIZE_BYTES)
                ),
            )
        return _download_document_by_ranges(
            source,
            download_url,
            session=session,
            headers=request_headers,
            verify=verify,
            chunk_size=int(request_config.get("range_chunk_size", _RANGE_FETCH_CHUNK_SIZE_BYTES)),
        )
    response = _get_with_retries(session, download_url, headers=request_headers, verify=verify)
    if _needs_browser_fallback(source, response):
        response.close()
        headers = {str(key): str(value) for key, value in session.headers.items()}
        headers["User-Agent"] = OFFICIAL_DOCUMENT_BROWSER_USER_AGENT
        response = _get_with_retries(session, download_url, headers=headers, verify=verify)
    if _needs_browser_fallback(source, response) and request_config.get("browser_impersonation"):
        response.close()
        impersonation_config = request_config.get("browser_impersonation")
        impersonate = (
            OFFICIAL_DOCUMENT_BROWSER_IMPERSONATION
            if impersonation_config is True
            else str(impersonation_config)
        )
        return _download_document_by_browser_impersonation(
            source,
            download_url,
            headers={**(request_headers or {}), "User-Agent": OFFICIAL_DOCUMENT_BROWSER_USER_AGENT},
            verify=verify,
            impersonate=impersonate,
        )
    if _needs_browser_fallback(source, response):
        response.close()
        raise RuntimeError(f"official document remained access-blocked: {download_url}")
    response.raise_for_status()
    return _DownloadedDocument(
        source=source,
        content=response.content,
        content_type=response.headers.get("content-type"),
        final_url=response.url,
    )


def _download_document_by_browser_impersonation(
    source: OfficialDocumentSource,
    download_url: str,
    *,
    headers: dict[str, str] | None,
    verify: bool,
    impersonate: str,
) -> _DownloadedDocument:
    """Fetch a document through curl_cffi for sources blocked by TLS fingerprinting."""
    try:
        from curl_cffi import requests as curl_requests
    except ImportError as exc:  # pragma: no cover - exercised only in incomplete installs
        raise RuntimeError("browser_impersonation official-document fetches require curl-cffi") from exc

    request_headers = {
        "User-Agent": OFFICIAL_DOCUMENT_BROWSER_USER_AGENT,
        **(headers or {}),
    }
    for attempt in range(1, _REQUEST_RETRY_ATTEMPTS + 1):
        try:
            response = curl_requests.get(
                download_url,
                headers=request_headers,
                timeout=90,
                allow_redirects=True,
                verify=verify,
                impersonate=cast(Any, impersonate),
            )
            if (
                response.status_code in _REQUEST_RETRY_STATUSES
                and attempt < _REQUEST_RETRY_ATTEMPTS
            ):
                cast(Any, response).close()
                _sleep_before_retry(attempt)
                continue
            cast(Any, response).raise_for_status()
            if _needs_browser_fallback(source, cast(requests.Response, response)):
                cast(Any, response).close()
                if attempt < _REQUEST_RETRY_ATTEMPTS:
                    _sleep_before_retry(attempt)
                    continue
                raise RuntimeError(
                    f"official document remained access-blocked after browser "
                    f"impersonation: {download_url}"
                )
            return _DownloadedDocument(
                source=source,
                content=response.content,
                content_type=response.headers.get("content-type"),
                final_url=str(response.url),
            )
        except Exception:
            if attempt >= _REQUEST_RETRY_ATTEMPTS:
                raise
            _sleep_before_retry(attempt)
    raise RuntimeError(f"failed to fetch official document with browser impersonation: {download_url}")


def _download_document_by_ranges(
    source: OfficialDocumentSource,
    download_url: str,
    *,
    session: requests.Session,
    headers: dict[str, str] | None,
    verify: bool,
    chunk_size: int,
) -> _DownloadedDocument:
    """Fetch a document through HTTP Range requests for servers that stall full GETs."""
    if chunk_size <= 0:
        raise ValueError("range_chunk_size must be positive")
    chunks: list[bytes] = []
    start = 0
    total: int | None = None
    content_type: str | None = None
    final_url = download_url
    while total is None or start < total:
        end = start + chunk_size - 1
        response = _get_with_retries(
            session,
            download_url,
            headers={**(headers or {}), "Range": f"bytes={start}-{end}"},
            verify=verify,
            stream=True,
        )
        response.raise_for_status()
        content_type = response.headers.get("content-type") or content_type
        final_url = response.url
        response_content = b"".join(response.iter_content(chunk_size=min(65536, chunk_size)))
        if response.status_code == 200:
            if chunks:
                response.close()
                raise RuntimeError(
                    f"range fetch for {source.source_id} returned full response after partial chunks"
                )
            return _DownloadedDocument(
                source=source,
                content=response_content,
                content_type=content_type,
                final_url=final_url,
            )
        if response.status_code != 206:
            response.close()
            raise RuntimeError(
                f"range fetch for {source.source_id} returned HTTP {response.status_code}"
            )
        range_start, range_end, range_total = _parse_content_range(
            str(response.headers.get("content-range") or "")
        )
        if range_start != start:
            response.close()
            raise RuntimeError(
                f"range fetch for {source.source_id} returned unexpected start {range_start}"
            )
        chunks.append(response_content)
        response.close()
        start = range_end + 1
        if range_total is not None:
            total = range_total
        if not chunks[-1] and total is None:
            break
    return _DownloadedDocument(
        source=source,
        content=b"".join(chunks),
        content_type=content_type,
        final_url=final_url,
    )


def _download_document_by_curl_ranges(
    source: OfficialDocumentSource,
    download_url: str,
    *,
    headers: dict[str, str] | None,
    verify: bool,
    chunk_size: int,
) -> _DownloadedDocument:
    """Fetch a document through curl Range requests for servers that stall urllib3."""
    if chunk_size <= 0:
        raise ValueError("range_chunk_size must be positive")
    chunks: list[bytes] = []
    start = 0
    total: int | None = None
    content_type: str | None = None
    request_headers = headers or {}
    user_agent = request_headers.get("User-Agent", OFFICIAL_DOCUMENT_USER_AGENT)
    with tempfile.TemporaryDirectory() as tmpdir_text:
        tmpdir = Path(tmpdir_text)
        while total is None or start < total:
            end = start + chunk_size - 1
            header_path = tmpdir / "headers.txt"
            body_path = tmpdir / "body.bin"
            command = [
                "curl",
                "-L",
                "--fail",
                "--silent",
                "--show-error",
                "--connect-timeout",
                "10",
                "--max-time",
                "60",
                "-A",
                user_agent,
            ]
            if not verify:
                command.append("--insecure")
            for key, value in request_headers.items():
                if key.lower() == "user-agent":
                    continue
                command.extend(["-H", f"{key}: {value}"])
            command.extend(
                [
                    "-H",
                    f"Range: bytes={start}-{end}",
                    "--dump-header",
                    str(header_path),
                    "--output",
                    str(body_path),
                    download_url,
                ]
            )
            subprocess.run(command, check=True)
            status_code, response_headers = _parse_curl_header_dump(
                header_path.read_text(errors="replace")
            )
            response_content = body_path.read_bytes()
            content_type = response_headers.get("content-type") or content_type
            if status_code == 200:
                if chunks:
                    raise RuntimeError(
                        f"curl range fetch for {source.source_id} returned full response after partial chunks"
                    )
                return _DownloadedDocument(
                    source=source,
                    content=response_content,
                    content_type=content_type,
                    final_url=download_url,
                )
            if status_code != 206:
                raise RuntimeError(
                    f"curl range fetch for {source.source_id} returned HTTP {status_code}"
                )
            range_start, range_end, range_total = _parse_content_range(
                response_headers.get("content-range", "")
            )
            if range_start != start:
                raise RuntimeError(
                    f"curl range fetch for {source.source_id} returned unexpected start {range_start}"
                )
            chunks.append(response_content)
            start = range_end + 1
            if range_total is not None:
                total = range_total
            if not response_content and total is None:
                break
    return _DownloadedDocument(
        source=source,
        content=b"".join(chunks),
        content_type=content_type,
        final_url=download_url,
    )


def _parse_content_range(header: str) -> tuple[int, int, int | None]:
    match = _CONTENT_RANGE_RE.match(header.strip())
    if not match:
        raise RuntimeError(f"invalid Content-Range header: {header!r}")
    total_text = match.group("total")
    return (
        int(match.group("start")),
        int(match.group("end")),
        None if total_text == "*" else int(total_text),
    )


def _parse_curl_header_dump(header_dump: str) -> tuple[int, dict[str, str]]:
    blocks = [
        block
        for block in re.split(r"(?:\r?\n){2,}", header_dump.strip())
        if block.lower().startswith("http/")
    ]
    if not blocks:
        raise RuntimeError("curl response did not include HTTP headers")
    lines = blocks[-1].splitlines()
    status_parts = lines[0].split()
    if len(status_parts) < 2:
        raise RuntimeError(f"invalid curl HTTP status line: {lines[0]!r}")
    headers: dict[str, str] = {}
    for line in lines[1:]:
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        headers[key.strip().lower()] = value.strip()
    return int(status_parts[1]), headers


def _request_headers_from_config(request_config: dict[str, Any]) -> dict[str, str] | None:
    if not request_config.get("browser_user_agent"):
        return None
    return {"User-Agent": OFFICIAL_DOCUMENT_BROWSER_USER_AGENT}


def _needs_browser_fallback(
    source: OfficialDocumentSource,
    response: requests.Response,
) -> bool:
    if response.status_code in _BROWSER_FALLBACK_STATUSES:
        return True
    declared_format = (source.source_format or "").lower()
    content_type = (response.headers.get("content-type") or "").lower()
    stripped = response.content.lstrip()
    if declared_format == "pdf" and not response.content.startswith(b"%PDF"):
        return "html" in content_type or stripped.startswith((b"<!doctype", b"<html"))
    return declared_format == "pdf" and _pdf_is_access_denial(response.content)


def _pdf_is_access_denial(content: bytes) -> bool:
    if not content.startswith(b"%PDF"):
        return False
    try:
        with fitz.open(stream=content, filetype="pdf") as document:
            if document.page_count > 2:
                return False
            text = " ".join(
                " ".join(page.get_text().split()) for page in document
            ).strip().lower()
    except (RuntimeError, ValueError):
        return False
    return len(text) <= 500 and text.startswith("the request is blocked.")


def _get_with_retries(
    session: requests.Session,
    url: str,
    *,
    headers: dict[str, str] | None = None,
    verify: bool = True,
    stream: bool = False,
) -> requests.Response:
    """Fetch a small official document, retrying transient server/network failures."""
    for attempt in range(1, _REQUEST_RETRY_ATTEMPTS + 1):
        try:
            with warnings.catch_warnings():
                if not verify:
                    warnings.simplefilter("ignore", InsecureRequestWarning)
                kwargs: dict[str, Any] = {
                    "headers": headers,
                    "timeout": 90,
                    "allow_redirects": True,
                    "verify": verify,
                }
                if stream:
                    kwargs["stream"] = True
                response = session.get(url, **kwargs)
            if (
                response.status_code in _REQUEST_RETRY_STATUSES
                and attempt < _REQUEST_RETRY_ATTEMPTS
            ):
                response.close()
                _sleep_before_retry(attempt)
                continue
            return response
        except requests.RequestException:
            if attempt >= _REQUEST_RETRY_ATTEMPTS:
                raise
            _sleep_before_retry(attempt)
    raise RuntimeError(f"failed to fetch official document: {url}")


def _sleep_before_retry(attempt: int) -> None:
    time.sleep(_REQUEST_RETRY_BASE_DELAY_SECONDS * attempt)


def _infer_source_format(source: OfficialDocumentSource, downloaded: _DownloadedDocument) -> str:
    if source.source_format:
        return source.source_format.lower()
    content_type = (downloaded.content_type or "").lower()
    if downloaded.content.startswith(b"%PDF") or "pdf" in content_type:
        return "pdf"
    if "html" in content_type or downloaded.content.lstrip().startswith((b"<!doctype", b"<html")):
        return "html"
    if "javascript" in content_type or source.source_url.lower().split("?", 1)[0].endswith(".js"):
        return "javascript"
    if "excel" in content_type or source.source_url.lower().split("?", 1)[0].endswith(".xls"):
        return "xls"
    if "msword" in content_type or downloaded.content.startswith(_LEGACY_WORD_DOCUMENT_MAGIC):
        return "doc"
    if (
        "wordprocessingml" in content_type
        or downloaded.content.startswith(b"PK")
        and _zip_contains(downloaded.content, "word/document.xml")
    ):
        return "docx"
    if (
        "spreadsheetml" in content_type
        or downloaded.content.startswith(b"PK")
        and _zip_contains(downloaded.content, "xl/workbook.xml")
    ):
        return "xlsx"
    raise ValueError(f"cannot infer source format for {source.source_id}")


def _sanitize_official_document_content(content: bytes, source_format: str) -> bytes:
    normalized_format = source_format.lower()
    if normalized_format not in {"html", "json"}:
        return content
    sanitized = _MAPBOX_PUBLIC_TOKEN_RE.sub(_MAPBOX_PUBLIC_TOKEN_PLACEHOLDER, content)
    if normalized_format != "json":
        return sanitized
    try:
        data = json_loads(sanitized.decode("utf-8-sig"))
    except (UnicodeDecodeError, ValueError):
        return sanitized
    return (json_dumps(data, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def _extension(source_format: str) -> str:
    if source_format == "pdf":
        return ".pdf"
    if source_format == "html":
        return ".html"
    if source_format == "javascript":
        return ".js"
    return f".{safe_segment(source_format)}"


def _extract_blocks(
    content: bytes,
    source_format: str,
    *,
    source_url: str,
    title: str | None,
    extraction: dict[str, Any] | None,
) -> tuple[_DocumentBlock, ...]:
    if source_format == "pdf":
        return _extract_pdf_blocks(content, extraction=extraction)
    if source_format == "html":
        return _extract_html_blocks(
            content, source_url=source_url, fallback_title=title, extraction=extraction
        )
    if source_format == "json":
        return _extract_json_html_blocks(
            content,
            source_url=source_url,
            fallback_title=title,
            extraction=extraction,
        )
    if source_format == "javascript":
        return _extract_plain_text_blocks(content, title=title)
    if source_format == "docx":
        return _extract_docx_blocks(content, extraction=extraction)
    if source_format == "doc":
        return _extract_doc_blocks(content, title=title, extraction=extraction)
    if source_format == "xlsx":
        return _extract_xlsx_blocks(content, extraction=extraction)
    if source_format == "xls":
        return _extract_xls_blocks(content, extraction=extraction)
    raise ValueError(f"unsupported official document source_format: {source_format}")


def _extract_plain_text_blocks(
    content: bytes,
    *,
    title: str | None,
) -> tuple[_DocumentBlock, ...]:
    """Extract a retained plain-text source as one auditable block."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1252")
    body = _normalize_text(text)
    if not body:
        return ()
    return (
        _DocumentBlock(
            kind="block",
            ordinal=1,
            heading=title,
            body=body,
            metadata={},
        ),
    )


def _extract_pdf_blocks(
    content: bytes, *, extraction: dict[str, Any] | None
) -> tuple[_DocumentBlock, ...]:
    extraction_config = extraction or {}
    segmentation = extraction_config.get("segmentation")
    if segmentation == "numbered_sections":
        return _extract_numbered_pdf_section_blocks(content, extraction=extraction_config)
    if segmentation == "labeled_sections":
        return _extract_labeled_pdf_section_blocks(content, extraction=extraction_config)
    if segmentation == "single_block":
        return _extract_single_block_pdf(content, extraction=extraction_config)
    blocks: list[_DocumentBlock] = []
    page_citation_prefix = extraction_config.get("page_citation_prefix")
    with fitz.open(stream=content, filetype="pdf") as document:
        for index, page in enumerate(document, start=1):
            text = _normalize_text(_pdf_page_text(page, extraction=extraction_config))
            if not text:
                continue
            metadata: dict[str, Any] = {"page_number": index}
            if page_citation_prefix:
                metadata["citation_suffix"] = f"{safe_segment(str(page_citation_prefix))}-{index}"
            blocks.append(
                _DocumentBlock(
                    kind="page",
                    ordinal=len(blocks) + 1,
                    heading=f"Page {index}",
                    body=text,
                    metadata=metadata,
                )
            )
    return tuple(blocks)


def _extract_single_block_pdf(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[_DocumentBlock, ...]:
    """Extract a whole PDF as one root provision (no per-page fragments).

    Use for short, logically indivisible documents (e.g. a single-schedule
    amending Act) where a page-per-provision split would both scatter the
    rule across fragments and grow the ``page-N`` citation-path ratchet. All
    page texts are concatenated in order under the source's own
    ``citation_path``; a blank-line separator preserves page boundaries for
    readers without emitting page suffixes.
    """
    page_texts: list[str] = []
    with fitz.open(stream=content, filetype="pdf") as document:
        for page in document:
            text = _normalize_text(_pdf_page_text(page, extraction=extraction))
            if text:
                page_texts.append(text)
    body = "\n\n".join(page_texts)
    if not body:
        return ()
    return (
        _DocumentBlock(
            kind="document",
            ordinal=1,
            heading=None,
            body=body,
            metadata={"page_count": len(page_texts)},
        ),
    )


def _extract_numbered_pdf_section_blocks(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[_DocumentBlock, ...]:
    """Extract legal PDFs whose top-level sections begin with numbered headings."""
    lines = _filtered_pdf_lines(content, extraction=extraction)
    sections: list[_DocumentBlock] = []
    index = 0
    while index < len(lines):
        match = _NUMBERED_SECTION_START_RE.match(lines[index][0])
        if not match:
            index += 1
            continue

        section_label = match.group("label")
        end_label = match.group("end_label")
        index += 1
        heading_text = match.group("heading")
        heading_lines: list[str] = [heading_text] if heading_text else []
        while index < len(lines) and _looks_like_section_heading_line(lines[index][0]):
            heading_lines.append(lines[index][0])
            index += 1
        if not heading_lines:
            continue

        body_lines: list[str] = []
        first_page = lines[index - 1][1]
        while index < len(lines) and not _NUMBERED_SECTION_START_RE.match(lines[index][0]):
            body_lines.append(lines[index][0])
            index += 1

        label_text = section_label if end_label is None else f"{section_label} -- {end_label}"
        citation_suffix = section_label if end_label is None else f"{section_label}-{end_label}"
        heading = f"{label_text}. {' '.join(heading_lines)}"
        page_numbers = [page for _line, page in lines[max(0, index - len(body_lines)) : index]]
        pages = page_numbers or [first_page]
        metadata: dict[str, Any] = {
            "citation_suffix": citation_suffix,
            "section_label": section_label,
            "page_start": min(pages),
            "page_end": max(pages),
        }
        if end_label is not None:
            metadata["section_end_label"] = end_label
        sections.append(
            _DocumentBlock(
                kind="section",
                ordinal=len(sections) + 1,
                heading=heading,
                body=_normalize_text("\n".join(body_lines)),
                metadata=metadata,
            )
        )
    return tuple(sections)


def _extract_labeled_pdf_section_blocks(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[_DocumentBlock, ...]:
    """Extract PDFs whose top-level sections begin with stable labels."""
    heading_pattern = extraction.get("section_heading_pattern")
    label_pattern = extraction.get("section_label_pattern")
    if heading_pattern is None and label_pattern is None:
        raise ValueError(
            "labeled_sections extraction requires section_heading_pattern or section_label_pattern"
        )
    section_heading_re = re.compile(str(heading_pattern)) if heading_pattern is not None else None
    section_label_re = re.compile(str(label_pattern)) if label_pattern is not None else None
    label_heading_pattern = extraction.get("label_only_heading_pattern")
    label_heading_re = (
        re.compile(str(label_heading_pattern)) if label_heading_pattern is not None else None
    )
    label_template = extraction.get("section_label_template")
    label_replacements = _section_label_replacements(extraction)
    label_requires_heading = bool(extraction.get("label_only_requires_heading", False))
    lines = _filtered_pdf_lines(content, extraction=extraction)
    drop_repeated = bool(extraction.get("drop_repeated_section_headings", True))
    heading_requires_bold = bool(extraction.get("section_heading_requires_bold", False))
    line_boldness = _pdf_line_boldness(content) if heading_requires_bold else {}
    line_occurrences: dict[tuple[str, int], int] = {}

    sections: list[_DocumentBlock] = []
    current_label: str | None = None
    current_heading: str | None = None
    current_body: list[str] = []
    current_body_pages: list[int] = []
    current_start_page: int | None = None
    index = 0

    def flush() -> None:
        nonlocal current_label, current_heading, current_body, current_body_pages
        nonlocal current_start_page
        if current_label is None or current_heading is None or current_start_page is None:
            return
        pages = current_body_pages or [current_start_page]
        sections.append(
            _DocumentBlock(
                kind="section",
                ordinal=len(sections) + 1,
                heading=current_heading,
                body=_normalize_text("\n".join(current_body)),
                metadata={
                    "citation_suffix": current_label,
                    "section_label": current_label,
                    "page_start": min(pages),
                    "page_end": max(pages),
                },
            )
        )
        current_label = None
        current_heading = None
        current_body = []
        current_body_pages = []
        current_start_page = None

    while index < len(lines):
        line, page = lines[index]
        line_location = (line, page)
        occurrence = line_occurrences.get(line_location, 0)
        line_occurrences[line_location] = occurrence + 1
        match = _match_labeled_pdf_section(
            line,
            section_heading_re,
            section_label_re,
            label_template=str(label_template) if label_template is not None else None,
            label_replacements=label_replacements,
        )
        if match and heading_requires_bold:
            occurrences = line_boldness.get(line_location, ())
            if occurrence >= len(occurrences) or not occurrences[occurrence]:
                match = None
        if match:
            label, heading_text = match
            consumed_label_heading = False
            if drop_repeated and label == current_label:
                index += 1
                while index < len(lines) and _looks_like_labeled_heading_continuation(
                    lines[index][0],
                    section_heading_re,
                    section_label_re,
                    label_template=str(label_template) if label_template is not None else None,
                    label_replacements=label_replacements,
                ):
                    index += 1
                continue
            if not heading_text and label_heading_re is not None:
                if index + 1 < len(lines) and label_heading_re.match(lines[index + 1][0]):
                    heading_text = lines[index + 1][0]
                    consumed_label_heading = True
                elif label_requires_heading:
                    if current_label is not None:
                        current_body.append(line)
                        current_body_pages.append(page)
                    index += 1
                    continue
            flush()
            heading_lines = [heading_text] if heading_text else []
            index += 1
            if consumed_label_heading:
                index += 1
            while index < len(lines) and _looks_like_labeled_heading_continuation(
                lines[index][0],
                section_heading_re,
                section_label_re,
                label_template=str(label_template) if label_template is not None else None,
                label_replacements=label_replacements,
            ):
                heading_lines.append(lines[index][0])
                index += 1
            heading = " ".join(part for part in heading_lines if part)
            current_label = label
            current_heading = f"{label} {heading}".strip()
            current_start_page = page
            continue
        if current_label is not None:
            current_body.append(line)
            current_body_pages.append(page)
        index += 1
    flush()
    return tuple(sections)


def _pdf_line_boldness(content: bytes) -> dict[tuple[str, int], tuple[bool, ...]]:
    """Return first-span boldness for each occurrence of a normalized PDF line."""
    line_styles: dict[tuple[str, int], list[bool]] = {}
    with fitz.open(stream=content, filetype="pdf") as document:
        for page_index, page in enumerate(document, start=1):
            for block in page.get_text("dict").get("blocks", ()):
                for line in block.get("lines", ()):
                    spans = line.get("spans", ())
                    first_span = next(
                        (span for span in spans if str(span.get("text", "")).strip()),
                        None,
                    )
                    if first_span is None:
                        continue
                    text = _normalize_text("".join(str(span.get("text", "")) for span in spans))
                    if text:
                        line_styles.setdefault((text, page_index), []).append(
                            bool(int(first_span.get("flags", 0)) & fitz.TEXT_FONT_BOLD)
                        )
    return {location: tuple(styles) for location, styles in line_styles.items()}


def _extract_docx_blocks(
    content: bytes, *, extraction: dict[str, Any] | None
) -> tuple[_DocumentBlock, ...]:
    if (extraction or {}).get("segmentation") == "labeled_sections":
        return _extract_labeled_docx_section_blocks(
            content,
            extraction=extraction or {},
        )
    with zipfile.ZipFile(BytesIO(content)) as document:
        xml = document.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    body = root.find("w:body", _WORD_NS)
    if body is None:
        return ()

    blocks: list[_DocumentBlock] = []
    heading: str | None = None
    parts: list[str] = []

    def flush() -> None:
        nonlocal parts
        body_text = _normalize_text("\n\n".join(parts))
        if body_text:
            blocks.append(
                _DocumentBlock(
                    kind="block",
                    ordinal=len(blocks) + 1,
                    heading=heading,
                    body=body_text,
                    metadata={},
                )
            )
        parts = []

    for child in body:
        if child.tag == _word_tag("p"):
            text = _docx_paragraph_text(child)
            if not text:
                continue
            if _docx_paragraph_is_heading(child):
                flush()
                heading = text
            else:
                parts.append(text)
        elif child.tag == _word_tag("tbl"):
            table_text = _docx_table_text(child)
            if table_text:
                parts.append(table_text)
    flush()
    return tuple(blocks)


def _extract_doc_blocks(
    content: bytes, *, title: str | None, extraction: dict[str, Any] | None
) -> tuple[_DocumentBlock, ...]:
    if (extraction or {}).get("segmentation") not in {None, "single_block"}:
        raise ValueError("legacy DOC extraction only supports single_block segmentation")
    body = _legacy_word_document_text(content)
    if not body:
        return ()
    heading = str((extraction or {}).get("heading") or title or "Document")
    return (
        _DocumentBlock(
            kind="block",
            ordinal=1,
            heading=heading,
            body=body,
            metadata={},
        ),
    )


def _extract_xlsx_blocks(
    content: bytes, *, extraction: dict[str, Any] | None
) -> tuple[_DocumentBlock, ...]:
    """Extract selected spreadsheet rows into corpus text blocks."""
    config = extraction or {}
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    try:
        sheet_names = _xlsx_sheet_names(workbook.sheetnames, config)
        blocks: list[_DocumentBlock] = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            header_row_number = int(config.get("xlsx_header_row") or config.get("header_row") or 1)
            start_row_number = int(
                config.get("xlsx_start_row") or config.get("start_row") or header_row_number + 1
            )
            headers: tuple[str, ...] | None = None
            output_columns = _xlsx_configured_strings(
                config.get("xlsx_columns") or config.get("columns")
            )
            filters = _xlsx_filters(config.get("xlsx_filters") or config.get("filters"))
            max_rows = config.get("xlsx_max_rows") or config.get("max_rows")
            row_limit = int(max_rows) if max_rows is not None else None
            selected_rows: list[tuple[int, tuple[str, ...]]] = []

            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_number == header_row_number:
                    headers = _xlsx_headers(row)
                    continue
                if row_number < start_row_number:
                    continue
                if headers is None:
                    headers = _xlsx_default_headers(len(row))
                index = _xlsx_header_index(headers)
                if not _xlsx_row_matches_filters(row, index=index, filters=filters):
                    continue
                row_columns = output_columns or headers
                selected_rows.append(
                    (
                        row_number,
                        tuple(
                            _xlsx_cell_text(row[index[column]])
                            if column in index and index[column] < len(row)
                            else ""
                            for column in row_columns
                        ),
                    )
                )
                if row_limit is not None and len(selected_rows) >= row_limit:
                    break

            if not selected_rows:
                continue
            row_columns = output_columns or headers or ()
            body_lines = [
                f"Sheet: {sheet_name}",
                "Row | " + " | ".join(row_columns),
            ]
            body_lines.extend(
                f"{row_number} | " + " | ".join(values) for row_number, values in selected_rows
            )
            blocks.append(
                _DocumentBlock(
                    kind="sheet",
                    ordinal=len(blocks) + 1,
                    heading=str(config.get("heading") or sheet_name),
                    body=_normalize_text("\n".join(body_lines)),
                    metadata=_spreadsheet_block_metadata(config, sheet_name, len(selected_rows)),
                )
            )
        return tuple(blocks)
    finally:
        workbook.close()


def _extract_xls_blocks(
    content: bytes, *, extraction: dict[str, Any] | None
) -> tuple[_DocumentBlock, ...]:
    """Extract selected legacy Excel workbook rows into corpus text blocks."""
    config = extraction or {}
    workbook = xlrd.open_workbook(file_contents=content)
    sheet_names = _xlsx_sheet_names(workbook.sheet_names(), config)
    blocks: list[_DocumentBlock] = []
    for sheet_name in sheet_names:
        worksheet = workbook.sheet_by_name(sheet_name)
        header_row_number = int(config.get("xls_header_row") or config.get("header_row") or 1)
        start_row_number = int(
            config.get("xls_start_row") or config.get("start_row") or header_row_number + 1
        )
        headers: tuple[str, ...] | None = None
        output_columns = _xlsx_configured_strings(
            config.get("xls_columns") or config.get("columns")
        )
        filters = _xlsx_filters(config.get("xls_filters") or config.get("filters"))
        max_rows = config.get("xls_max_rows") or config.get("max_rows")
        row_limit = int(max_rows) if max_rows is not None else None
        selected_rows: list[tuple[int, tuple[str, ...]]] = []

        for row_number in range(1, worksheet.nrows + 1):
            row = tuple(
                worksheet.cell_value(row_number - 1, column) for column in range(worksheet.ncols)
            )
            if row_number == header_row_number:
                headers = _xlsx_headers(row)
                continue
            if row_number < start_row_number:
                continue
            if headers is None:
                headers = _xlsx_default_headers(len(row))
            index = _xlsx_header_index(headers)
            if not _xlsx_row_matches_filters(row, index=index, filters=filters):
                continue
            row_columns = output_columns or headers
            selected_rows.append(
                (
                    row_number,
                    tuple(
                        _xlsx_cell_text(row[index[column]])
                        if column in index and index[column] < len(row)
                        else ""
                        for column in row_columns
                    ),
                )
            )
            if row_limit is not None and len(selected_rows) >= row_limit:
                break

        if not selected_rows:
            continue
        row_columns = output_columns or headers or ()
        body_lines = [
            f"Sheet: {sheet_name}",
            "Row | " + " | ".join(row_columns),
        ]
        body_lines.extend(
            f"{row_number} | " + " | ".join(values) for row_number, values in selected_rows
        )
        blocks.append(
            _DocumentBlock(
                kind="sheet",
                ordinal=len(blocks) + 1,
                heading=str(config.get("heading") or sheet_name),
                body=_normalize_text("\n".join(body_lines)),
                metadata=_spreadsheet_block_metadata(config, sheet_name, len(selected_rows)),
            )
        )
    return tuple(blocks)


def _spreadsheet_block_metadata(
    config: dict[str, Any], sheet_name: str, row_count: int
) -> dict[str, Any]:
    metadata = {
        "sheet_name": sheet_name,
        "row_count": row_count,
    }
    citation_suffix = config.get("citation_suffix") or config.get("section_label")
    if isinstance(citation_suffix, str) and citation_suffix:
        metadata["citation_suffix"] = citation_suffix
        metadata["section_label"] = citation_suffix
    return metadata


def _xlsx_sheet_names(available_sheet_names: list[str], config: dict[str, Any]) -> tuple[str, ...]:
    raw_sheets = config.get("xlsx_sheets") or config.get("sheets")
    raw_sheet = config.get("xlsx_sheet") or config.get("sheet")
    if raw_sheets is None and raw_sheet is not None:
        raw_sheets = (raw_sheet,)
    if raw_sheets is None:
        return tuple(available_sheet_names)
    configured = _xlsx_configured_strings(raw_sheets)
    missing = [sheet for sheet in configured if sheet not in available_sheet_names]
    if missing:
        raise ValueError(f"xlsx sheet not found: {', '.join(missing)}")
    return configured


def _xlsx_configured_strings(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    if isinstance(value, str):
        return (value,)
    if isinstance(value, (list, tuple)):
        return tuple(str(item) for item in value)
    raise ValueError("xlsx configuration value must be a string or list of strings")


def _xlsx_filters(raw_filters: Any) -> dict[str, tuple[str, ...]]:
    if raw_filters is None:
        return {}
    if not isinstance(raw_filters, dict):
        raise ValueError("xlsx filters must be a mapping")
    filters: dict[str, tuple[str, ...]] = {}
    for key, value in raw_filters.items():
        if isinstance(value, (list, tuple)):
            filters[str(key)] = tuple(_xlsx_cell_text(item) for item in value)
        else:
            filters[str(key)] = (_xlsx_cell_text(value),)
    return filters


def _xlsx_headers(row: tuple[Any, ...]) -> tuple[str, ...]:
    headers = tuple(
        _xlsx_cell_text(value) or f"column_{index}" for index, value in enumerate(row, start=1)
    )
    return _dedupe_xlsx_headers(headers)


def _xlsx_default_headers(width: int) -> tuple[str, ...]:
    return tuple(f"column_{index}" for index in range(1, width + 1))


def _dedupe_xlsx_headers(headers: tuple[str, ...]) -> tuple[str, ...]:
    counts: dict[str, int] = {}
    deduped: list[str] = []
    for header in headers:
        count = counts.get(header, 0) + 1
        counts[header] = count
        deduped.append(header if count == 1 else f"{header}_{count}")
    return tuple(deduped)


def _xlsx_header_index(headers: tuple[str, ...]) -> dict[str, int]:
    return {header: index for index, header in enumerate(headers)}


def _xlsx_row_matches_filters(
    row: tuple[Any, ...],
    *,
    index: dict[str, int],
    filters: dict[str, tuple[str, ...]],
) -> bool:
    for column, allowed_values in filters.items():
        if column not in index:
            raise ValueError(f"xlsx filter column not found: {column}")
        cell_index = index[column]
        actual = _xlsx_cell_text(row[cell_index]) if cell_index < len(row) else ""
        if actual not in allowed_values:
            return False
    return True


def _xlsx_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat) and not isinstance(value, (int, float, str)):
        return str(isoformat())
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return _normalize_text(str(value))


def _legacy_word_document_text(content: bytes) -> str:
    """Extract text from legacy binary Word documents using available system tools."""
    with tempfile.TemporaryDirectory(prefix="axiom-doc-") as temp_dir:
        doc_path = Path(temp_dir) / "source.doc"
        doc_path.write_bytes(content)

        command: list[str] | None = None
        if shutil.which("textutil"):
            command = ["textutil", "-convert", "txt", "-stdout", str(doc_path)]
        elif shutil.which("antiword"):
            command = ["antiword", str(doc_path)]
        elif shutil.which("catdoc"):
            command = ["catdoc", str(doc_path)]
        if command is None:
            raise RuntimeError("legacy DOC extraction requires textutil, antiword, or catdoc")

        result = subprocess.run(
            command,
            check=False,
            capture_output=True,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            message = result.stderr.strip() or result.stdout.strip()
            raise RuntimeError(f"legacy DOC extraction failed: {message}")
        return _normalize_text(result.stdout)


def _extract_labeled_docx_section_blocks(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[_DocumentBlock, ...]:
    """Extract DOCX documents whose provisions begin with stable labels."""
    heading_pattern = extraction.get("section_heading_pattern")
    label_pattern = extraction.get("section_label_pattern")
    if heading_pattern is None and label_pattern is None:
        raise ValueError(
            "labeled_sections DOCX extraction requires section_heading_pattern "
            "or section_label_pattern"
        )
    section_heading_re = re.compile(str(heading_pattern)) if heading_pattern is not None else None
    section_label_re = re.compile(str(label_pattern)) if label_pattern is not None else None
    label_template = extraction.get("section_label_template")
    label_replacements = _section_label_replacements(extraction)
    start_after_pattern = extraction.get("start_after_pattern")
    start_after_re = (
        re.compile(str(start_after_pattern)) if start_after_pattern is not None else None
    )
    stop_pattern = extraction.get("stop_text_pattern")
    stop_re = re.compile(str(stop_pattern)) if stop_pattern is not None else None
    drop_lines = {str(line).strip() for line in extraction.get("drop_lines", ())}
    drop_line_patterns = tuple(
        re.compile(str(pattern)) for pattern in extraction.get("drop_line_patterns", ())
    )

    sections: list[_DocumentBlock] = []
    current_label: str | None = None
    current_heading: str | None = None
    current_body: list[str] = []
    started = start_after_re is None

    def flush() -> None:
        nonlocal current_label, current_heading, current_body
        if current_label is None:
            return
        sections.append(
            _DocumentBlock(
                kind="section",
                ordinal=len(sections) + 1,
                heading=f"{current_label} {current_heading or ''}".strip(),
                body=_normalize_text("\n\n".join(current_body)),
                metadata={
                    "citation_suffix": current_label,
                    "section_label": current_label,
                },
            )
        )
        current_label = None
        current_heading = None
        current_body = []

    for line in _docx_lines(content):
        if not started:
            if start_after_re is not None and start_after_re.search(line):
                started = True
            continue
        if _drop_pdf_line(line, drop_lines, drop_line_patterns):
            continue
        if stop_re is not None and stop_re.search(line):
            flush()
            break
        match = _match_labeled_pdf_section(
            line,
            section_heading_re,
            section_label_re,
            label_template=str(label_template) if label_template is not None else None,
            label_replacements=label_replacements,
        )
        if match:
            label, heading = match
            flush()
            current_label = label
            current_heading = heading or label
            current_body = []
            continue
        if current_label is not None:
            current_body.append(line)
    flush()
    return tuple(sections)


def _docx_lines(content: bytes) -> tuple[str, ...]:
    with zipfile.ZipFile(BytesIO(content)) as document:
        xml = document.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    body = root.find("w:body", _WORD_NS)
    if body is None:
        return ()
    lines: list[str] = []
    for child in body:
        if child.tag == _word_tag("p"):
            text = _docx_paragraph_text(child)
            if text:
                lines.append(text)
        elif child.tag == _word_tag("tbl"):
            table_text = _docx_table_text(child)
            if table_text:
                lines.extend(line for line in table_text.splitlines() if line)
    return tuple(lines)


def _zip_contains(content: bytes, name: str) -> bool:
    try:
        with zipfile.ZipFile(BytesIO(content)) as archive:
            return name in archive.namelist()
    except zipfile.BadZipFile:
        return False


def _word_tag(local_name: str) -> str:
    return f"{{{_WORD_NS['w']}}}{local_name}"


def _docx_paragraph_is_heading(paragraph: ElementTree.Element) -> bool:
    style = paragraph.find("w:pPr/w:pStyle", _WORD_NS)
    value = style.get(_word_tag("val")) if style is not None else None
    if not value:
        return False
    normalized = value.lower().replace(" ", "")
    return normalized.startswith("heading") or normalized in {"title", "subtitle"}


def _docx_paragraph_text(paragraph: ElementTree.Element) -> str:
    return _normalize_text("".join(_docx_text_chunks(paragraph)))


def _docx_table_text(table: ElementTree.Element) -> str:
    rows: list[str] = []
    for row in table.findall("w:tr", _WORD_NS):
        cells = [
            _normalize_text(" ".join(_docx_text_chunks(cell)))
            for cell in row.findall("w:tc", _WORD_NS)
        ]
        cells = [cell for cell in cells if cell]
        if cells:
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def _docx_text_chunks(node: ElementTree.Element) -> tuple[str, ...]:
    chunks: list[str] = []
    for descendant in node.iter():
        if descendant.tag == _word_tag("t") and descendant.text:
            chunks.append(descendant.text)
        elif descendant.tag == _word_tag("tab"):
            chunks.append("\t")
        elif descendant.tag == _word_tag("br"):
            chunks.append("\n")
    return tuple(chunks)


def _match_labeled_pdf_section(
    line: str,
    section_heading_re: re.Pattern[str] | None,
    section_label_re: re.Pattern[str] | None,
    *,
    label_template: str | None = None,
    label_replacements: dict[str, str] | None = None,
) -> tuple[str, str] | None:
    if section_heading_re is not None:
        match = section_heading_re.match(line)
        if match:
            heading_text = match.groupdict().get("heading") or ""
            return (
                _labeled_section_label(
                    match,
                    label_template=label_template,
                    label_replacements=label_replacements,
                ),
                heading_text.strip(),
            )
    if section_label_re is not None:
        match = section_label_re.match(line)
        if match:
            return (
                _labeled_section_label(
                    match,
                    label_template=label_template,
                    label_replacements=label_replacements,
                ),
                "",
            )
    return None


def _filtered_pdf_lines(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[tuple[str, int], ...]:
    if extraction.get("page_windows") is not None:
        return _windowed_pdf_lines(content, extraction=extraction)
    start_page = _positive_int(extraction.get("start_page"), default=1)
    end_page = extraction.get("end_page")
    parsed_end_page = _positive_int(end_page, default=0) if end_page is not None else None
    if parsed_end_page is not None and parsed_end_page < start_page:
        raise ValueError("end_page must be greater than or equal to start_page")
    drop_lines = {str(line).strip() for line in extraction.get("drop_lines", ())}
    drop_line_patterns = tuple(
        re.compile(str(pattern)) for pattern in extraction.get("drop_line_patterns", ())
    )
    start_after_pattern = extraction.get("start_after_pattern")
    start_after_re = (
        re.compile(str(start_after_pattern)) if start_after_pattern is not None else None
    )
    started = start_after_re is None
    lines: list[tuple[str, int]] = []
    with fitz.open(stream=content, filetype="pdf") as document:
        for page_index, page in enumerate(document, start=1):
            if page_index < start_page:
                continue
            if parsed_end_page is not None and page_index > parsed_end_page:
                break
            for raw_line in _pdf_page_text(page, extraction=extraction).splitlines():
                line = _normalize_text(raw_line)
                if not line or _drop_pdf_line(line, drop_lines, drop_line_patterns):
                    continue
                if not started:
                    if start_after_re is not None and start_after_re.search(line):
                        started = True
                    continue
                lines.append((line, page_index))
    return tuple(lines)


@dataclass(frozen=True)
class _PdfPageWindow:
    start_page: int
    end_page: int
    start_at_re: re.Pattern[str] | None
    stop_at_re: re.Pattern[str] | None


def _parse_pdf_page_windows(extraction: dict[str, Any]) -> tuple[_PdfPageWindow, ...]:
    for legacy_key in ("start_page", "end_page", "start_after_pattern"):
        if extraction.get(legacy_key) is not None:
            raise ValueError(f"page_windows cannot be combined with {legacy_key}")
    raw_windows = extraction.get("page_windows")
    if not isinstance(raw_windows, (list, tuple)) or not raw_windows:
        raise ValueError("page_windows must be a non-empty list of window mappings")
    windows: list[_PdfPageWindow] = []
    previous_end = 0
    for raw in raw_windows:
        if not isinstance(raw, dict):
            raise ValueError("each page window must be a mapping")
        start_page = _positive_int(raw.get("start_page"), default=0)
        end_page = _positive_int(raw.get("end_page"), default=0)
        if not start_page or not end_page:
            raise ValueError("each page window requires start_page and end_page")
        if end_page < start_page:
            raise ValueError("page window end_page must be >= start_page")
        if start_page <= previous_end:
            raise ValueError("page windows must be ascending and non-overlapping")
        previous_end = end_page
        start_at = raw.get("start_at_pattern")
        stop_at = raw.get("stop_at_pattern")
        windows.append(
            _PdfPageWindow(
                start_page=start_page,
                end_page=end_page,
                start_at_re=re.compile(str(start_at)) if start_at is not None else None,
                stop_at_re=re.compile(str(stop_at)) if stop_at is not None else None,
            )
        )
    return tuple(windows)


def _windowed_pdf_lines(
    content: bytes, *, extraction: dict[str, Any]
) -> tuple[tuple[str, int], ...]:
    """Collect PDF text lines from discontiguous page windows.

    Each window is a mapping with ``start_page``/``end_page`` (1-based,
    inclusive) plus optional line anchors: ``start_at_pattern`` drops lines
    until the first matching line (the matching line is kept, so a section
    label can anchor the window), and ``stop_at_pattern`` ends the window at
    the first matching line (the matching line is dropped). Windows let one
    manifest entry capture discontiguous sections of a large statute PDF —
    one source snapshot instead of one per slice — while keeping stray
    neighbour-section fragments out of the extracted bodies.
    """
    windows = _parse_pdf_page_windows(extraction)
    drop_lines = {str(line).strip() for line in extraction.get("drop_lines", ())}
    drop_line_patterns = tuple(
        re.compile(str(pattern)) for pattern in extraction.get("drop_line_patterns", ())
    )
    window_by_page: dict[int, _PdfPageWindow] = {}
    for window in windows:
        for page_number in range(window.start_page, window.end_page + 1):
            window_by_page[page_number] = window
    last_page = max(window.end_page for window in windows)
    started: dict[int, bool] = {id(window): window.start_at_re is None for window in windows}
    stopped: dict[int, bool] = {id(window): False for window in windows}
    lines: list[tuple[str, int]] = []
    with fitz.open(stream=content, filetype="pdf") as document:
        for page_index, page in enumerate(document, start=1):
            if page_index > last_page:
                break
            page_window = window_by_page.get(page_index)
            if page_window is None or stopped[id(page_window)]:
                continue
            for raw_line in _pdf_page_text(page, extraction=extraction).splitlines():
                line = _normalize_text(raw_line)
                if not line or _drop_pdf_line(line, drop_lines, drop_line_patterns):
                    continue
                if not started[id(page_window)]:
                    if page_window.start_at_re is not None and page_window.start_at_re.search(
                        line
                    ):
                        started[id(page_window)] = True
                        lines.append((line, page_index))
                    continue
                if page_window.stop_at_re is not None and page_window.stop_at_re.search(line):
                    stopped[id(page_window)] = True
                    break
                lines.append((line, page_index))
    return tuple(lines)


def _pdf_page_text(page: Any, *, extraction: dict[str, Any]) -> str:
    if extraction.get("force_ocr"):
        return _ocr_pdf_page_text(page, extraction=extraction)
    text = page.get_text("text", sort=bool(extraction.get("sort_text")))
    if _normalize_text(text) or not extraction.get("ocr"):
        return str(text)
    return _ocr_pdf_page_text(page, extraction=extraction)


def _ocr_pdf_page_text(page: Any, *, extraction: dict[str, Any]) -> str:
    """Extract text from an image-only PDF page using the local Tesseract CLI."""
    if not shutil.which("tesseract"):
        raise RuntimeError("PDF OCR extraction requires tesseract on PATH")

    dpi = _positive_int(extraction.get("ocr_dpi"), default=200)
    language = str(extraction.get("ocr_language") or "eng")
    page_segmentation_mode = extraction.get("ocr_psm")

    with tempfile.TemporaryDirectory(prefix="axiom-pdf-ocr-") as temp_dir:
        image_path = Path(temp_dir) / "page.png"
        # Render with PyMuPDF's target-DPI rasterizer rather than an
        # equivalent zoom matrix. On some scanned legal PDFs the matrix path
        # anti-aliases faint interior table cells into illegibility (e.g. the
        # Ghana Act 1111 rate table dropped its 10%/17.5% band cells), while
        # the dpi= rasterizer preserves them, yielding faithful OCR.
        pixmap = page.get_pixmap(dpi=dpi, alpha=False)
        pixmap.save(str(image_path))

        command = ["tesseract", str(image_path), "stdout", "-l", language]
        if page_segmentation_mode is not None:
            command.extend(["--psm", str(page_segmentation_mode)])
        result = subprocess.run(
            command,
            check=False,
            capture_output=True,
            encoding="utf-8",
            errors="replace",
        )
        if result.returncode != 0:
            message = result.stderr.strip() or result.stdout.strip()
            raise RuntimeError(f"PDF OCR extraction failed: {message}")
        return result.stdout


_NUMBERED_SECTION_START_RE = re.compile(
    r"^(?P<label>\d{3})\.(?:\s*--\s*(?P<end_label>\d{3})\.)?(?:\s+(?P<heading>.+))?$"
)


def _positive_int(value: Any, *, default: int) -> int:
    if value is None:
        return default
    parsed = int(value)
    if parsed < 1:
        raise ValueError("page numbers must be positive")
    return parsed


def _drop_pdf_line(
    line: str, drop_lines: set[str], drop_line_patterns: tuple[re.Pattern[str], ...]
) -> bool:
    if line in drop_lines:
        return True
    if re.match(r"^Page \d+$", line):
        return True
    if re.match(r"^Section \d{3}$", line):
        return True
    return any(pattern.search(line) for pattern in drop_line_patterns)


def _looks_like_section_heading_line(line: str) -> bool:
    if line in {"(RESERVED)", "(RESERVED)."}:
        return True
    if line in _NON_HEADING_UPPERCASE_LINES:
        return False
    letters = [character for character in line if character.isalpha()]
    if not letters:
        return False
    uppercase_letters = [character for character in letters if character.isupper()]
    return len(uppercase_letters) / len(letters) >= 0.75


def _looks_like_labeled_heading_continuation(
    line: str,
    section_heading_re: re.Pattern[str] | None,
    section_label_re: re.Pattern[str] | None,
    *,
    label_template: str | None = None,
    label_replacements: dict[str, str] | None = None,
) -> bool:
    if _match_labeled_pdf_section(
        line,
        section_heading_re,
        section_label_re,
        label_template=label_template,
        label_replacements=label_replacements,
    ):
        return False
    return _looks_like_section_heading_line(line)


def _extract_html_blocks(
    content: bytes,
    *,
    source_url: str,
    fallback_title: str | None,
    extraction: dict[str, Any] | None,
) -> tuple[_DocumentBlock, ...]:
    soup = _html_soup(content)
    drop_selectors = [
        "script",
        "style",
        "noscript",
        "svg",
        "button",
        "input",
        "nav",
        "select",
        "header",
        "footer",
        "textarea",
        "aside",
        ".breadcrumb",
        ".breadcrumbs",
        "[aria-label='breadcrumb']",
        *_html_drop_selectors(extraction),
    ]
    for selector in drop_selectors:
        for node in soup.select(selector):
            node.decompose()
    root = _html_content_root(soup, extraction=extraction)
    title = _document_title(soup) or fallback_title
    if (extraction or {}).get("segmentation") == "anchor_range":
        return _extract_anchor_range_html_blocks(
            root,
            title=title,
            source_url=source_url,
            extraction=extraction or {},
        )
    webworks_blocks = _extract_webworks_html_blocks(root, title=title, source_url=source_url)
    if webworks_blocks:
        return webworks_blocks
    if (extraction or {}).get("segmentation") == "labeled_sections":
        return _extract_labeled_html_section_blocks(
            root,
            title=title,
            source_url=source_url,
            extraction=extraction or {},
        )
    blocks: list[_DocumentBlock] = []
    heading = title
    parts: list[str] = []

    def flush() -> None:
        nonlocal parts
        body = _normalize_text("\n\n".join(parts))
        if body:
            blocks.append(
                _DocumentBlock(
                    kind="block",
                    ordinal=len(blocks) + 1,
                    heading=heading,
                    body=body,
                    metadata={"source_url": source_url},
                )
            )
        parts = []

    for node in root.find_all(_TEXT_TAGS):
        if not isinstance(node, Tag) or _inside_text_tag(node):
            continue
        text = _normalize_text(node.get_text(" ", strip=True))
        if not text:
            continue
        if node.name in _HEADING_TAGS:
            flush()
            heading = text
            continue
        parts.append(text)
    flush()
    if blocks:
        return tuple(blocks)
    fallback = _normalize_text(root.get_text(" ", strip=True))
    if not fallback:
        return ()
    return (
        _DocumentBlock(
            kind="block",
            ordinal=1,
            heading=title,
            body=fallback,
            metadata={"source_url": source_url},
        ),
    )


def _extract_json_html_blocks(
    content: bytes,
    *,
    source_url: str,
    fallback_title: str | None,
    extraction: dict[str, Any] | None,
) -> tuple[_DocumentBlock, ...]:
    if (extraction or {}).get("segmentation") == "source_only":
        json_loads(content.decode("utf-8-sig"))
        return ()
    if (extraction or {}).get("segmentation") == "records":
        return _extract_json_record_blocks(
            content,
            source_url=source_url,
            fallback_title=fallback_title,
            extraction=extraction or {},
        )
    html_field = (extraction or {}).get("json_html_field")
    if not isinstance(html_field, str) or not html_field:
        raise ValueError(
            "json official document extraction requires json_html_field or segmentation: records"
        )
    data = json_loads(content.decode("utf-8"))
    html_text = _json_path(data, html_field)
    if not isinstance(html_text, str) or not html_text.strip():
        raise ValueError(f"json_html_field did not resolve to HTML text: {html_field}")
    if (extraction or {}).get("json_html_base64"):
        html_text = base64.b64decode(html_text).decode("utf-8", errors="replace")
    if (extraction or {}).get("json_html_as_single_block"):
        soup = BeautifulSoup(html_text, "html.parser")
        body = _normalize_text(soup.get_text(" ", strip=True))
        if not body:
            return ()
        return (
            _DocumentBlock(
                kind="block",
                ordinal=1,
                heading=fallback_title,
                body=body,
                metadata={"source_url": source_url},
            ),
        )
    return _extract_html_blocks(
        html_text.encode("utf-8"),
        source_url=source_url,
        fallback_title=fallback_title,
        extraction=extraction,
    )


def _extract_json_record_blocks(
    content: bytes,
    *,
    source_url: str,
    fallback_title: str | None,
    extraction: dict[str, Any],
) -> tuple[_DocumentBlock, ...]:
    """Extract structured JSON API records with one HTML/text body per record."""
    records_path = extraction.get("json_records_path")
    text_field = extraction.get("json_record_text_field")
    label_field = extraction.get("json_record_label_field")
    citation_suffix_field = extraction.get("json_record_citation_suffix_field")
    heading_field = extraction.get("json_record_heading_field")
    kind_field = extraction.get("json_record_kind_field")
    status_field = extraction.get("json_record_status_field")
    include_statuses = extraction.get("json_record_include_statuses")
    exclude_statuses = extraction.get("json_record_exclude_statuses", ())
    metadata_fields = extraction.get("json_record_metadata_fields", ())
    text_is_html = bool(extraction.get("json_record_text_is_html", True))
    citation_suffix_slugify = bool(extraction.get("json_record_citation_suffix_slugify", False))

    if not isinstance(text_field, str) or not text_field:
        raise ValueError("records JSON extraction requires json_record_text_field")
    if label_field is not None and not isinstance(label_field, str):
        raise ValueError("json_record_label_field must be a string when configured")
    if citation_suffix_field is not None and not isinstance(citation_suffix_field, str):
        raise ValueError("json_record_citation_suffix_field must be a string when configured")
    if heading_field is not None and not isinstance(heading_field, str):
        raise ValueError("json_record_heading_field must be a string when configured")
    if kind_field is not None and not isinstance(kind_field, str):
        raise ValueError("json_record_kind_field must be a string when configured")
    if status_field is not None and not isinstance(status_field, str):
        raise ValueError("json_record_status_field must be a string when configured")
    if isinstance(include_statuses, str):
        include_statuses = (include_statuses,)
    if isinstance(exclude_statuses, str):
        exclude_statuses = (exclude_statuses,)
    if isinstance(metadata_fields, str):
        metadata_fields = (metadata_fields,)
    include_status_set = {str(status) for status in include_statuses or ()}
    exclude_status_set = {str(status) for status in exclude_statuses or ()}
    metadata_field_names = tuple(str(field) for field in metadata_fields)

    data = json_loads(content.decode("utf-8"))
    rows = _json_path(data, str(records_path)) if records_path else data
    if not isinstance(rows, list):
        raise ValueError("json_records_path must resolve to a list")

    blocks: list[_DocumentBlock] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        status = _json_record_value(row, status_field)
        if include_status_set and str(status) not in include_status_set:
            continue
        if exclude_status_set and str(status) in exclude_status_set:
            continue

        raw_text_value = _json_record_value(row, text_field)
        raw_text = _json_record_text(raw_text_value)
        if not raw_text:
            continue
        body = _html_fragment_text(raw_text) if text_is_html else _normalize_text(raw_text)
        if not body:
            continue

        label_value = _json_record_value(row, label_field)
        citation_suffix_value = _json_record_value(row, citation_suffix_field)
        heading_value = _json_record_value(row, heading_field)
        kind_value = _json_record_value(row, kind_field)
        label = str(label_value).strip() if label_value not in {None, ""} else ""
        citation_suffix = (
            str(citation_suffix_value).strip()
            if citation_suffix_value not in {None, ""}
            else label
        )
        if citation_suffix_slugify:
            citation_suffix = _json_record_citation_suffix_slug(citation_suffix)
        heading_text = (
            str(heading_value).strip()
            if heading_value not in {None, ""}
            else fallback_title or label or "Record"
        )
        heading = f"{label} {heading_text}".strip() if label else heading_text
        metadata = {
            "source_url": source_url,
            **{
                field: row.get(field)
                for field in metadata_field_names
                if field in row and field != text_field
            },
        }
        if citation_suffix:
            metadata["citation_suffix"] = citation_suffix
        if label:
            metadata["section_label"] = label
        elif citation_suffix:
            metadata["section_label"] = citation_suffix
        blocks.append(
            _DocumentBlock(
                kind=str(kind_value or "record").strip().lower(),
                ordinal=len(blocks) + 1,
                heading=heading,
                body=body,
                metadata=metadata,
            )
        )
    return tuple(blocks)


def _json_record_text(value: Any) -> str:
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, list):
        parts = [str(item).strip() for item in value if item is not None and str(item).strip()]
        return "\n\n".join(parts)
    return ""


def _json_record_citation_suffix_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")


def _html_soup(content: bytes) -> BeautifulSoup:
    """Parse official HTML with a parser that preserves malformed void tags."""
    try:
        return BeautifulSoup(content, "lxml")
    except FeatureNotFound:
        return BeautifulSoup(content, "html.parser")


def _json_path(data: Any, path: str) -> Any:
    current = data
    for part in path.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
            continue
        raise ValueError(f"json path did not resolve: {path}")
    return current


def _json_record_value(row: dict[str, Any], field: str | None) -> Any:
    if not field:
        return None
    current: Any = row
    for part in field.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
            continue
        return None
    return current


def _html_fragment_text(html_text: str) -> str:
    soup = BeautifulSoup(html_text, "html.parser")
    for comment in soup.find_all(string=lambda node: isinstance(node, Comment)):
        comment.extract()
    for selector in ("script", "style", "noscript", "svg"):
        for node in soup.select(selector):
            node.decompose()
    return _normalize_text(soup.get_text(" ", strip=True))


def _extract_labeled_html_section_blocks(
    root: Tag,
    *,
    title: str | None,
    source_url: str,
    extraction: dict[str, Any],
) -> tuple[_DocumentBlock, ...]:
    """Extract HTML documents whose provisions begin with stable text labels."""
    heading_pattern = extraction.get("section_heading_pattern")
    label_pattern = extraction.get("section_label_pattern")
    if heading_pattern is None and label_pattern is None:
        raise ValueError(
            "labeled_sections HTML extraction requires section_heading_pattern "
            "or section_label_pattern"
        )
    section_heading_re = re.compile(str(heading_pattern)) if heading_pattern is not None else None
    section_label_re = re.compile(str(label_pattern)) if label_pattern is not None else None
    label_template = extraction.get("section_label_template")
    label_replacements = _section_label_replacements(extraction)
    stop_pattern = extraction.get("stop_text_pattern")
    stop_re = re.compile(str(stop_pattern)) if stop_pattern is not None else None

    sections: list[_DocumentBlock] = []
    current_label: str | None = None
    current_heading: str | None = None
    current_body: list[str] = []

    def flush() -> None:
        nonlocal current_label, current_heading, current_body
        if current_label is None:
            return
        heading = current_heading or title or current_label
        sections.append(
            _DocumentBlock(
                kind="section",
                ordinal=len(sections) + 1,
                heading=heading,
                body=_normalize_text("\n\n".join(current_body)),
                metadata={
                    "citation_suffix": current_label,
                    "section_label": current_label,
                    "source_url": source_url,
                },
            )
        )
        current_label = None
        current_heading = None
        current_body = []

    for node in root.find_all(_TEXT_TAGS):
        if not isinstance(node, Tag) or _inside_text_tag(node):
            continue
        text = _normalize_text(node.get_text(" ", strip=True))
        if not text:
            continue
        if stop_re is not None and stop_re.match(text):
            flush()
            break
        match = _match_labeled_html_section(
            text,
            section_heading_re,
            section_label_re,
            label_template=str(label_template) if label_template is not None else None,
            label_replacements=label_replacements,
        )
        if match is not None:
            label, heading, body = match
            flush()
            current_label = label
            current_heading = heading or label
            current_body = [body] if body else []
            continue
        if current_label is not None:
            current_body.append(text)
    flush()
    return tuple(sections)


def _extract_anchor_range_html_blocks(
    root: Tag,
    *,
    title: str | None,
    source_url: str,
    extraction: dict[str, Any],
) -> tuple[_DocumentBlock, ...]:
    """Extract one HTML section from a start node through the next configured node."""
    ranges = extraction.get("anchor_ranges")
    if ranges is not None:
        if not isinstance(ranges, list):
            raise ValueError("anchor_ranges must be a list of mappings")
        shared_config = {key: value for key, value in extraction.items() if key != "anchor_ranges"}
        blocks: list[_DocumentBlock] = []
        for range_config in ranges:
            if not isinstance(range_config, dict):
                raise ValueError("anchor_ranges entries must be mappings")
            block = _extract_anchor_range_html_block(
                root,
                title=title,
                source_url=source_url,
                extraction={**shared_config, **range_config},
                ordinal=len(blocks) + 1,
            )
            if block is not None:
                blocks.append(block)
        return tuple(blocks)
    block = _extract_anchor_range_html_block(
        root,
        title=title,
        source_url=source_url,
        extraction=extraction,
        ordinal=1,
    )
    return (block,) if block is not None else ()


def _extract_anchor_range_html_block(
    root: Tag,
    *,
    title: str | None,
    source_url: str,
    extraction: dict[str, Any],
    ordinal: int,
) -> _DocumentBlock | None:
    start_selector = extraction.get("html_start_selector") or extraction.get("start_selector")
    if not isinstance(start_selector, str) or not start_selector:
        raise ValueError("anchor_range HTML extraction requires html_start_selector")
    start = root.select_one(start_selector)
    if not isinstance(start, Tag):
        raise ValueError(f"html start selector did not match: {start_selector!r}")

    stop_selector = extraction.get("html_stop_selector") or extraction.get("stop_selector")
    stop = None
    if stop_selector is not None:
        if not isinstance(stop_selector, str) or not stop_selector:
            raise ValueError("html_stop_selector must be a non-empty string")
        stop = root.select_one(stop_selector)
        if not isinstance(stop, Tag):
            raise ValueError(f"html stop selector did not match: {stop_selector!r}")

    html_parts: list[str] = []
    for node in (start, *start.next_siblings):
        if isinstance(node, Comment):
            continue
        if stop is not None and _html_node_contains(node, stop):
            break
        html_parts.append(str(node))

    text = _html_fragment_text("".join(html_parts))
    if not text:
        return None
    label = extraction.get("section_label") or extraction.get("citation_suffix")
    heading = extraction.get("section_heading") or title or label
    metadata = {"source_url": source_url}
    if isinstance(label, str) and label:
        metadata["citation_suffix"] = label
        metadata["section_label"] = label
    return _DocumentBlock(
        kind="section",
        ordinal=ordinal,
        heading=str(heading) if heading else None,
        body=text,
        metadata=metadata,
    )


def _html_node_contains(node: Any, target: Tag) -> bool:
    if node is target:
        return True
    return isinstance(node, Tag) and any(descendant is target for descendant in node.descendants)


def _match_labeled_html_section(
    text: str,
    section_heading_re: re.Pattern[str] | None,
    section_label_re: re.Pattern[str] | None,
    *,
    label_template: str | None = None,
    label_replacements: dict[str, str] | None = None,
) -> tuple[str, str, str] | None:
    if section_heading_re is not None:
        match = section_heading_re.match(text)
        if match:
            groups = match.groupdict()
            label = _labeled_section_label(
                match,
                label_template=label_template,
                label_replacements=label_replacements,
            )
            return (
                label,
                (groups.get("heading") or "").strip(),
                (groups.get("body") or "").strip(),
            )
    if section_label_re is not None:
        match = section_label_re.match(text)
        if match:
            return (
                _labeled_section_label(
                    match,
                    label_template=label_template,
                    label_replacements=label_replacements,
                ),
                "",
                "",
            )
    return None


def _labeled_section_label(
    match: re.Match[str],
    *,
    label_template: str | None,
    label_replacements: dict[str, str] | None = None,
) -> str:
    groups = {key: (value or "").strip() for key, value in match.groupdict().items()}
    if label_template:
        try:
            return _replace_section_label(
                label_template.format(**groups).strip(),
                label_replacements,
            )
        except KeyError as exc:
            raise ValueError(
                f"section_label_template references unknown group: {exc.args[0]}"
            ) from exc
    label = groups.get("label")
    if label:
        return _replace_section_label(label, label_replacements)
    raise ValueError(
        "labeled_sections extraction requires a label group unless "
        "section_label_template is configured"
    )


def _section_label_replacements(extraction: dict[str, Any]) -> dict[str, str] | None:
    raw = extraction.get("section_label_replacements")
    if raw is None:
        return None
    if not isinstance(raw, dict):
        raise ValueError("section_label_replacements must be a mapping")
    return {str(key): str(value).strip() for key, value in raw.items()}


def _replace_section_label(label: str, replacements: dict[str, str] | None) -> str:
    if not replacements:
        return label
    return replacements.get(label, label).strip()


def _html_content_root(soup: BeautifulSoup, *, extraction: dict[str, Any] | None) -> Tag:
    selector = (extraction or {}).get("html_content_selector") or (extraction or {}).get(
        "content_selector"
    )
    if selector is not None:
        root = soup.select_one(str(selector))
        if isinstance(root, Tag):
            return root
        raise ValueError(f"html content selector did not match: {selector!r}")
    return _main_content(soup)


def _html_drop_selectors(extraction: dict[str, Any] | None) -> tuple[str, ...]:
    selectors = (extraction or {}).get("html_drop_selectors") or (extraction or {}).get(
        "drop_selectors"
    )
    if selectors is None:
        return ()
    if isinstance(selectors, str):
        return (selectors,)
    return tuple(str(selector) for selector in selectors)


_WEBWORKS_TEXT_CLASS_PREFIXES = (
    "Body_Text",
    "List_",
    "Note_",
    "Numbered_",
    "Cell_",
    "Table_",
)


def _extract_webworks_html_blocks(
    root: Tag,
    *,
    title: str | None,
    source_url: str,
) -> tuple[_DocumentBlock, ...]:
    page = root.select_one("#page_content")
    if not isinstance(page, Tag):
        return ()
    for selector in (".WebWorks_MiniTOC", ".ww_skin_page_globalization"):
        for node in page.select(selector):
            node.decompose()

    blocks: list[_DocumentBlock] = []
    heading = title
    parts: list[str] = []

    def flush() -> None:
        nonlocal parts
        body = _normalize_text("\n\n".join(parts))
        if body:
            blocks.append(
                _DocumentBlock(
                    kind="block",
                    ordinal=len(blocks) + 1,
                    heading=heading,
                    body=body,
                    metadata={"source_url": source_url},
                )
            )
        parts = []

    for node in page.find_all(_is_webworks_content_node):
        text = _normalize_text(node.get_text(" ", strip=True))
        if not text:
            continue
        if _has_webworks_heading_class(node):
            flush()
            heading = text
            continue
        parts.append(text)
    flush()
    return tuple(blocks)


def _is_webworks_content_node(node: Tag) -> bool:
    return _has_webworks_heading_class(node) or any(
        _has_class_prefix(node, prefix) for prefix in _WEBWORKS_TEXT_CLASS_PREFIXES
    )


def _has_webworks_heading_class(node: Tag) -> bool:
    return _has_class_prefix(node, "Heading_")


def _has_class_prefix(node: Tag, prefix: str) -> bool:
    classes = node.get("class")
    if not isinstance(classes, list):
        return False
    return any(isinstance(item, str) and item.startswith(prefix) for item in classes)


def _main_content(soup: BeautifulSoup) -> Tag:
    for selector in ("main", "article", "[role='main']", "#main-content", ".main-content"):
        node = soup.select_one(selector)
        if isinstance(node, Tag):
            return node
    if isinstance(soup.body, Tag):
        return soup.body
    return soup


def _document_title(soup: BeautifulSoup) -> str | None:
    h1 = soup.find("h1")
    if isinstance(h1, Tag):
        text = _normalize_text(h1.get_text(" ", strip=True))
        if text:
            return text
    if soup.title:
        text = _normalize_text(soup.title.get_text(" ", strip=True))
        if text:
            return text
    return None


def _inside_text_tag(node: Tag) -> bool:
    for parent in node.parents:
        if not isinstance(parent, Tag):
            continue
        if parent.name in _TEXT_TAGS:
            return True
    return False


def _inventory_items(
    source: OfficialDocumentSource,
    *,
    blocks: tuple[_DocumentBlock, ...],
    source_key: str,
    source_format: str,
    source_sha: str,
    content_type: str | None,
    final_url: str,
) -> tuple[SourceInventoryItem, ...]:
    root_path = _root_citation_path(source)
    metadata = _source_metadata(
        source,
        content_type=content_type,
        final_url=final_url,
        block_count=len(blocks),
    )
    items = [
        SourceInventoryItem(
            citation_path=root_path,
            source_url=source.source_url,
            source_path=source_key,
            source_format=source_format,
            sha256=source_sha,
            metadata={"kind": "document", **metadata},
        )
    ]
    for block in blocks:
        items.append(
            SourceInventoryItem(
                citation_path=_block_citation_path(source, block),
                source_url=source.source_url,
                source_path=source_key,
                source_format=source_format,
                sha256=source_sha,
                metadata={"kind": block.kind, **metadata, **block.metadata},
            )
        )
    return tuple(items)


def _provision_records(
    source: OfficialDocumentSource,
    *,
    blocks: tuple[_DocumentBlock, ...],
    version: str,
    source_key: str,
    source_format: str,
    source_as_of: str,
    expression_date: str,
    content_type: str | None,
    final_url: str,
) -> tuple[ProvisionRecord, ...]:
    root_path = _root_citation_path(source)
    root_id = deterministic_provision_id(root_path)
    metadata = _source_metadata(
        source,
        content_type=content_type,
        final_url=final_url,
        block_count=len(blocks),
    )
    records = [
        ProvisionRecord(
            id=root_id,
            jurisdiction=source.jurisdiction,
            document_class=source.document_class,
            citation_path=root_path,
            heading=source.title,
            citation_label=source.title,
            version=version,
            source_url=source.source_url,
            source_path=source_key,
            source_id=source.source_id,
            source_format=source_format,
            source_as_of=source_as_of,
            expression_date=expression_date,
            language=source.language or "en",
            level=1,
            ordinal=1,
            kind="document",
            metadata={"kind": "document", **metadata},
        )
    ]
    for block in blocks:
        citation_path = _block_citation_path(source, block)
        records.append(
            ProvisionRecord(
                id=deterministic_provision_id(citation_path),
                jurisdiction=source.jurisdiction,
                document_class=source.document_class,
                citation_path=citation_path,
                body=block.body,
                heading=block.heading,
                citation_label=f"{source.title} {block.ordinal}",
                version=version,
                source_url=source.source_url,
                source_path=source_key,
                source_id=source.source_id,
                source_format=source_format,
                source_as_of=source_as_of,
                expression_date=expression_date,
                language=source.language or "en",
                parent_citation_path=root_path,
                parent_id=root_id,
                level=2,
                ordinal=block.ordinal,
                kind=block.kind,
                metadata={"kind": block.kind, **metadata, **block.metadata},
            )
        )
    return tuple(records)


def _source_metadata(
    source: OfficialDocumentSource,
    *,
    content_type: str | None,
    final_url: str,
    block_count: int,
) -> dict[str, Any]:
    metadata = dict(source.metadata or {})
    metadata.update(
        {
            "title": source.title,
            "content_type": content_type,
            "download_url": final_url,
            "block_count": block_count,
        }
    )
    return metadata


def _root_citation_path(source: OfficialDocumentSource) -> str:
    if source.citation_path:
        return _validate_citation_path(
            source.citation_path,
            jurisdiction=source.jurisdiction,
            document_class=source.document_class,
        )
    return f"{source.jurisdiction}/{source.document_class}/{safe_segment(source.source_id)}"


def _block_citation_path(source: OfficialDocumentSource, block: _DocumentBlock) -> str:
    citation_suffix = block.metadata.get("citation_suffix")
    if isinstance(citation_suffix, str) and citation_suffix:
        return f"{_root_citation_path(source)}/{safe_segment(citation_suffix)}"
    return f"{_root_citation_path(source)}/{block.kind}-{block.ordinal}"


def _validate_citation_path(
    citation_path: str,
    *,
    jurisdiction: str,
    document_class: str,
) -> str:
    """Return a manifest-supplied citation path after basic scope validation."""
    normalized = citation_path.strip().strip("/")
    expected_prefix = f"{jurisdiction}/{document_class}/"
    if not normalized.startswith(expected_prefix):
        raise ValueError(f"citation_path must start with {expected_prefix!r}: {citation_path!r}")
    for part in normalized.split("/"):
        safe_segment(part)
    return normalized


def _date_text(value: date | str | None, fallback: str) -> str:
    if value is None:
        return fallback
    if isinstance(value, date):
        return value.isoformat()
    return value


def _normalize_text(text: str) -> str:
    text = text.replace("\u200b", "").replace("\ufeff", "")
    lines = [" ".join(line.split()) for line in text.splitlines()]
    paragraphs: list[str] = []
    current: list[str] = []
    for line in lines:
        if not line:
            if current:
                paragraphs.append(" ".join(current))
                current = []
            continue
        current.append(line)
    if current:
        paragraphs.append(" ".join(current))
    return "\n\n".join(paragraphs)


def _progress(stream: TextIO | None, message: str) -> None:
    if stream is None:
        return
    print(message, file=stream)
    stream.flush()


if __name__ == "__main__":
    print(
        "Use `axiom-corpus-ingest extract-official-documents` to run this adapter.",
        file=sys.stderr,
    )
